"""
The configDao class is a data access object that performs basic CRUD
(create, read, update, delete) functions on the config database.
"""

import io
import os
import pkgutil
from typing import Union, List
from concurrent.futures import Future, ThreadPoolExecutor
import yaml
from dynamodb import DynamoDao
from config import WrfConfig
from openpyxl import Workbook
import openpyxl
from  system import get_aws_session


class ConfigDao(DynamoDao):
    """
    CRUD operations for configs
    """

    def __init__(self, endpoint_url: str = None):
        """
        Create the Data Access Object (DAO)
        :param endpoint_url: (optional) Specifying the endpoint URL can be useful for testing
        """
        # load the config table definition
        #self.table_definition = yaml.safe_load(pkgutil.get_data('wrfcloud', 'config/table.yaml'))

        # get the table name
        table_name = os.getenv('DOMAIN_DB')

        # get the key fields for the table
        key_fields = ['name']

        # get the endpoint URL, but do not overwrite the local argument
        if endpoint_url is None and 'ENDPOINT_URL' in os.environ:
            endpoint_url = os.environ['ENDPOINT_URL']
        # call the super constructor
        super().__init__(table_name, key_fields, endpoint_url)

    def add_config(self, config: WrfConfig) -> bool:
        """
        Store a new config
        :param config: config object to store
        :return: True if successful, otherwise False
        """
        # track the success state
        ok = True

        # deep copy the configuration data
        config_data = config.data
        #print('config_data')
        #print(config_data)
        # save the namelists to S3
        if 'wrf_namelist' in config_data:
            wrf_namelist: str = config_data.pop('wrf_namelist')
            ok = ok and self._save_namelist(config.s3_key_wrf_namelist, wrf_namelist)
        if 'wrf_bk_namelist' in config_data:
            wrf_bk_namelist: str = config_data.pop('wrf_bk_namelist')
            ok = ok and self._save_namelist(config.s3_key_wrf_bk_namelist, wrf_bk_namelist)
        if 'wps_namelist' in config_data:
            wps_namelist: str = config_data.pop('wps_namelist')
            ok = ok and self._save_namelist(config.s3_key_wps_namelist, wps_namelist)
        if 'location' in config_data:
            location: dict = config_data.pop('location')
            ok = ok and self._save_xlsx(config.s3_key_location, location)
        # save the item to the database
        return ok and super().put_item(config_data)

    def get_config_by_name(self, name: str) -> Union[WrfConfig, None]:
        """
        Get a config by config id
        :param name: config name
        :return: The config with the given name, or None if not found
        """
        # build the database key
        key = {'name': name}
        # get the item with the key
        data = super().get_item(key)
        # handle the case where the key is not found
        if data is None:
            return None

        # build a new config object
        config = WrfConfig(data)

        # load the namelists from S3
        
        self._load_namelist(config, config.s3_key_wrf_namelist)
        self._load_namelist(config, config.s3_key_wrf_bk_namelist)
        self._load_namelist(config, config.s3_key_wps_namelist)
        # load the location file from S3
        self._load_xlsx(config, config.s3_key_location)

        return config

    def get_all_configs(self) -> List[WrfConfig]:
        """
        Get a list of all configs in the system
        :return: List of all configs
        """
        # Convert a list of items into a list of User objects
        configs: List[WrfConfig] = [WrfConfig(item) for item in super().get_all_items()]

        # Load namelists from S3
        #tpe = ThreadPoolExecutor(max_workers=16)
        #futures: List[Future] = [tpe.submit(self._load_namelist, config, config.s3_key_wrf_namelist) for config in configs]
        #futures += [tpe.submit(self._load_namelist, config, config.s3_key_wps_namelist) for config in configs]
        #for future in futures:
        #    future.result()

        return configs

    def update_config(self, config: WrfConfig) -> bool:
        """
        Update the config data
        :param config: config data values to update, which must include the key field (config_id)
        :return: True if successful, otherwise False
        """
        # track the success state
        ok = True

        # deep copy the configuration data
        config_data = config.data

        # save the namelists to S3
        if 'wrf_namelist' in config_data:
            wrf_namelist: str = config_data.pop('wrf_namelist')
            ok = ok and self._save_namelist(config.s3_key_wrf_namelist, wrf_namelist)
        if 'wrf_bk_namelist' in config_data:
            wrf_bk_namelist: str = config_data.pop('wrf_bk_namelist')
            ok = ok and self._save_namelist(config.s3_key_wrf_bk_namelist, wrf_bk_namelist)
        if 'wps_namelist' in config_data:
            wps_namelist: str = config_data.pop('wps_namelist')
            ok = ok and self._save_namelist(config.s3_key_wps_namelist, wps_namelist)
        if 'location' in config_data:
            location: dict = config_data.pop('location')
            ok = ok and self._save_xlsx(config.s3_key_location, location)
        # save the item to the database
        return ok and super().update_item(config_data)

    def delete_config(self, config: WrfConfig) -> bool:
        """
        Delete the config from the database (not any associated data)
        :param config: WrfConfig object
        :return: True if successful, otherwise False
        """
        # delete the s3 objects
        ok = self._delete_namelist(config.s3_key_wrf_namelist)
        ok = ok and self._delete_namelist(config.s3_key_wps_namelist)
        ok = ok and self._delete_namelist(config.s3_key_wrf_bk_namelist)
        ok = ok and self._delete_namelist(config.s3_key_location)
        # delete the dynamodb item
        return ok and super().delete_item({'name': config.name})

    def create_config_table(self) -> bool:
        """
        Create the config table
        :return: True if successful, otherwise False
        """
        return super().create_table(
            self.table_definition['attribute_definitions'],
            self.table_definition['key_schema']
        )

    def _save_namelist(self, namelist_key: str, namelist_data: str) -> bool:
        """
        Save the namelist to S3
        :param namelist_key: S3 key for the namelist
        :param namelist_data: Contents of the namelist file
        :return: True if successfully saved to S3, otherwise False
        """
        data = io.BytesIO(namelist_data.encode())
        bucket = os.environ['WRFCLOUD_BUCKET']

        try:
            s3 = get_aws_session().client('s3')
            s3.upload_fileobj(data, bucket, namelist_key)
        except Exception as e:
            self.log.error('Failed to write namelist to S3.', e)
            return False

        return True
        
    def _save_xlsx(self, location_key: str, xlsx_data: dict) -> bool:
        """
        Save the namelist to S3
        :param namelist_key: S3 key for the namelist
        :param namelist_data: Contents of the namelist file
        :return: True if successfully saved to S3, otherwise False
        """
        #print(xlsx_data)
        #df = pd.DataFrame(xlsx_data[1:], columns=xlsx_data[0])
        #with io.BytesIO() as buffer:
        #    writer = pd.ExcelWriter(buffer, engine='xlsxwriter')
        #    df.to_excel(writer, index=False)
        #    writer._save()
        #    data = buffer.getvalue()
        wb=Workbook()
        ws=wb.active
        ws.title='locations'
        for row in xlsx_data:
            ws.append(row)
        file=io.BytesIO()
        wb.save(file)
        file.seek(0)

        bucket_name = os.environ['WRFCLOUD_BUCKET']

        try:
            s3 = get_aws_session().resource('s3')
            bucket=s3.Bucket(bucket_name)
            bucket.put_object(Key=location_key, Body=file)
        except Exception as e:
            self.log.error('Failed to write location file to S3.', e)
            return False

        return True

    def _load_namelist(self, config: WrfConfig, namelist_key: str) -> bool:
        """
        Load the namelist from S3
        :param config: WRF configuration object
        :param namelist_key: S3 key of the namelist to load
        :return: True if successfully loaded, otherwise False
        """
        # get the bucket name from the environment
        bucket = os.environ['WRFCLOUD_BUCKET']
  
        # load the namelist into the config data
        try:
            s3 = get_aws_session().client('s3')
            data: str = s3.get_object(Bucket=bucket, Key=namelist_key)['Body'].read().decode()

            if namelist_key.endswith('namelist.input'):
                config.wrf_namelist = data
            elif namelist_key.endswith('namelist.wps'):
                config.wps_namelist = data
            elif namelist_key.endswith('namelist.input.bk'):
                config.wrf_bk_namelist = data
        except Exception as e:
            self.log.error(f'Failed to read namelist from S3. s3://{bucket}/{namelist_key}', e)
            return False

    def _load_xlsx(self, config: WrfConfig, location_key: str) -> bool:
        # get the bucket name from the environment
        bucket = os.environ['WRFCLOUD_BUCKET']
        # load the namelist into the config data
        try:
            s3 = get_aws_session().client('s3')
            response = s3.get_object(Bucket=bucket, Key=location_key)
            # Load workbook from S3
            
            wb = openpyxl.load_workbook(io.BytesIO(response['Body'].read()))

            # Get the active worksheet
            ws = wb.active

            # Create an empty list
            data_list = []

            # Iterate through rows and columns
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                data_list.append([cell.value for cell in row])
            config.location = data_list
        except Exception as e:
            self.log.error(f'Failed to read location file from S3. s3://{bucket}/{location_key}', e)
            return False            
            
    def _delete_namelist(self, namelist_key: str) -> bool:
        """
        Delete the namelist from S3
        :param namelist_key: S3 key for the namelist
        :return: True if successful, otherwise False
        """
        bucket = os.environ['WRFCLOUD_BUCKET']
        try:
            s3 = get_aws_session().client('s3')
            s3.delete_object(Bucket=bucket, Key=namelist_key)
        except Exception as e:
            self.log.error('Failed to delete namelist from S3.', e)
            return False

        return True
